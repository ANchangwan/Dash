import dash
import dash_pivottable
import dash_html_components as html
from openpyxl.xml.constants import MAX_COLUMN
import pandas as pd
from dash.dependencies import Input, Output
import dash_core_components as dcc
import dash_html_components as html
from openpyxl import load_workbook
import plotly.graph_objects as go
import plotly.express as px



app = dash.Dash(__name__)
app.title = "Life Dream Data Analysis"

# 엑셀 데이터 분석 및 가공
df = pd.read_excel("C:/Users/com/Documents/Python Scripts/매출내역.xlsx")
df.columns = ["no","주문일","업체명","구분","매출액","입금액","입금일","세금계산","카드","현금영수증","담당번호","주문내역","받는분","주소", "문자발송 H.P","송장번호","구비서류","담당자","비고","비고2","비고3"]
df2 = pd.read_excel("C:/Users/com/Documents/Python Scripts/월별매출액.xlsx")
df2.columns = ["월","월별매출액"]
x = df["매출액"]


wb = load_workbook("C:/Users/com/Documents/Python Scripts/매출내역.xlsx")
ws = wb.active


# 엑셀 데이터 범위 지정
# data_range = ws[3:ws.max_row-12]
data_range = ws["B3:F320"]
data = [[cell.value for cell in rows] for rows in data_range ]


#  html 형식으로 만들기
app.layout = html.Div(
    dash_pivottable.PivotTable(
        data=data,
        rows = ["구분"],
        rowOrder="key_a_to_z",
        cols = ["업체명"],
        colOrder="key_a_to_z",
        rendererName="Grouped Column Chart",
        vals = ["매출액"]

    )
)

# 서버 
if __name__ == "__main__":
    app.run_server(debug = True)
